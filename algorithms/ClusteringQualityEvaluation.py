# Public evaluation helper for saved clustering labels.
from __future__ import annotations

from pathlib import Path
import os
import numpy as np
from openpyxl import load_workbook
from scipy.optimize import linear_sum_assignment
from sklearn.metrics import adjusted_rand_score, adjusted_mutual_info_score, confusion_matrix

REPO_ROOT = Path(__file__).resolve().parents[1]
RESULTS_ROOT = REPO_ROOT / "experiment outcomes"
RECORDS_PATH = REPO_ROOT / "experiment_records" / "GBSK run records.xlsx"
DATASET_NAME = "DryBean"
LABELS2_PATH = REPO_ROOT / "datasets" / DATASET_NAME / "labels.txt"
ALGORITHM_NAME = "GBSK"


def check_ami_in_log(log_path: Path) -> bool:
    try:
        last_line = log_path.read_text(encoding="utf-8", errors="ignore").splitlines()[-1].strip()
    except IndexError:
        return False
    return "AMI:" in last_line or "AMI" in last_line


def find_seed_row(sheet, seed):
    for row in range(2, sheet.max_row + 1):
        if sheet.cell(row=row, column=6).value == int(seed):
            return row
    return None


def main() -> None:
    main_dir = RESULTS_ROOT / ALGORITHM_NAME / DATASET_NAME
    if not main_dir.exists():
        raise FileNotFoundError(f"Results directory not found: {main_dir}")

    labels2 = np.loadtxt(LABELS2_PATH)
    wb = load_workbook(RECORDS_PATH)
    sheet = wb[DATASET_NAME]

    for folder_name in sorted(os.listdir(main_dir)):
        folder_path = main_dir / folder_name
        if not folder_path.is_dir():
            continue

        labels_file = folder_path / "labels.txt"
        log_file = folder_path / "log.txt"
        if not labels_file.exists() or not log_file.exists() or check_ami_in_log(log_file):
            continue

        seed = folder_name.split(' ')[0].replace('Seed_', '')
        labels1 = np.loadtxt(labels_file)
        cm = confusion_matrix(labels1, labels2)
        row_ind, col_ind = linear_sum_assignment(cm, maximize=True)
        accuracy = cm[row_ind, col_ind].sum() / np.sum(cm)
        ari = adjusted_rand_score(labels1, labels2)
        ami = adjusted_mutual_info_score(labels1, labels2)

        seed_row = find_seed_row(sheet, seed)
        if seed_row is not None:
            sheet.cell(row=seed_row, column=18, value=accuracy)
            sheet.cell(row=seed_row, column=19, value=ari)
            sheet.cell(row=seed_row, column=20, value=ami)

        with log_file.open("a", encoding="utf-8") as log:
            log.write(f"ACC: {accuracy:.4f}\n")
            log.write(f"ARI: {ari:.4f}\n")
            log.write(f"AMI: {ami:.4f}\n")

        print(f"{folder_name}: ACC={accuracy:.4f} ARI={ari:.4f} AMI={ami:.4f}")

    wb.save(RECORDS_PATH)


if __name__ == "__main__":
    main()
